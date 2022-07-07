import pandas as pd
import sys
import time
import os
import re

from os import system, name
from os.path import exists
from app.definitions import *
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


from rich.console import Console
from rich.prompt import Prompt
from rich.theme import Theme
from rich.table import Table
from rich.progress import track


# from rich import console.print

custom_theme = Theme(
    {
        "primary": "bold cyan",
        "success": "bold bright_green",
        "info": "cyan",
        "warning": "yellow",
        "danger": "bright_red",
        "url": "not bold not italic underline bright_blue",
    }
)

console = Console(theme=custom_theme)


def debugging():
    global debugMode
    debugMode = not debugMode
    return debugMode


def createBin(file_name: str, mode: str, key: str, desc: str):
    try:
        creation_name = (
            "/" + str(file_name[: len(file_name) - 5]).strip() + " (Bin)" + ".xlsx"
        )
        creation_path = "./bin/" + mode

        if not os.path.exists(creation_path):
            os.makedirs(creation_path)

        if not exists(creation_path + creation_name):
            writer = pd.ExcelWriter(creation_path + creation_name, engine="xlsxwriter")
            writer.save()
            book = load_workbook(creation_path + creation_name)
            sheet = book.active
            sheet.append(bin_header)
            book.save(creation_path + creation_name)

        book = load_workbook(creation_path + creation_name)
        sheet = book.active

        rowData = (key, desc)
        sheet.append(rowData)
        book.save(creation_path + creation_name)

        if debugMode:
            console.print(desc, style="warning")
    except Exception as e:
        if debugMode:
            console.print(":x: Error: " + str(e), style="danger")


def getMachineries():
    try:
        machineries: list = []

        path = "./data/name_list.xlsx"
        mach_list = pd.read_excel(path)

        i = 0
        while not pd.isna(mach_list.iloc[i, 1]) and mach_list.iloc[i, 1] != "END":
            machineries.append([str(mach_list.iloc[i, 0]), str(mach_list.iloc[i, 1])])
            i += 1

        return machineries
    except Exception as e:
        if debugMode:
            console.print(":x: Error: " + str(e), style="danger")


def getMachinery(
    machinery_id: str,
    key: str,
    mode: str,
    file_name: str,
    machineries: list,
):
    try:
        if not pd.isna(machinery_id) or machinery_id != "":
            machinery_id = machinery_id.strip()

        for machinery in machineries:
            if machinery[1] == machinery_id or machinery[1] == key:
                return str(machinery[0])

        createBin(
            file_name,
            mode,
            key,
            "⚠️ Warning: No machinery ( " + machinery_id + " ) found for " + key,
        )

        return "N/A"

    except Exception as e:
        if debugMode:
            console.print(
                ":x: Error: " + str(e) + " (" + key + ": " + machinery_id + ")"
            )


def getCodes():
    try:
        codes: list = []

        path = "./data/code_list.xlsx"
        code_list = pd.read_excel(path)

        i = 0
        while not pd.isna(code_list.iloc[i, 1]) and code_list.iloc[i, 1] != "END":
            codes.append([str(code_list.iloc[i, 0]), str(code_list.iloc[i, 1])])
            i += 1

        return codes
    except Exception as e:
        if debugMode:
            console.print(":x: Error: " + str(e), style="danger")


def getCode(
    machinery_name: str,
    key: str,
    mode: str,
    file_name: str,
    codes: list,
):
    try:
        if not pd.isna(machinery_name) or machinery_name != "":
            machinery_name = machinery_name.strip()

        for code in codes:
            if code[1] == machinery_name or code[1] == key:
                return str(code[0])

        createBin(
            file_name,
            mode,
            key,
            "⚠️ Warning: No machinery code ( " + machinery_name + " ) found for " + key,
        )

        return "N/A"
    except Exception as e:
        if debugMode:
            console.print(
                ":x: Error: " + str(e) + " (" + key + ": " + machinery_name + ")",
                style="warning",
            )


def getIntervals():
    try:
        intervals: list = []

        path = "./data/interval_list.xlsx"
        interval_list = pd.read_excel(path)

        i = 0
        while (
            not pd.isna(interval_list.iloc[i, 1]) and interval_list.iloc[i, 1] != "END"
        ):
            intervals.append(
                [str(interval_list.iloc[i, 0]), str(interval_list.iloc[i, 1])]
            )
            i += 1

        return intervals
    except Exception as e:
        if debugMode:
            console.print(":x: Error: " + str(e), style="danger")


def getInterval(
    interval_id: str,
    key: str,
    mode: str,
    file_name: str,
    intervals: list,
):
    try:
        if not pd.isna(interval_id) or interval_id != "":
            interval_id = interval_id.strip()

        for interval in intervals:
            if interval[1] == interval_id or interval[1] == key:
                return str(interval[0])

        createBin(
            file_name,
            mode,
            key,
            "⚠️ Warning: No interval ( " + interval_id + " ) found for " + key,
        )

        return "N/A"
    except Exception as e:
        if debugMode:
            console.print(
                ":x: Error: " + str(e) + " (" + key + ": " + interval_id + ")",
                style="warning",
            )


def has_numbers(inputString: str):
    try:
        return bool(re.search(r"\d", inputString))
    except Exception as e:
        if debugMode:
            console.print(":x: Error: " + str(e), style="danger")


def header():

    console.print(
        r"""
    ____              ______                     __         
   / __ \__  __      / ____/___  _________  ____/ /__  _____
  / /_/ / / / /_____/ __/ / __ \/ ___/ __ \/ __  / _ \/ ___/
 / ____/ /_/ /_____/ /___/ / / / /__/ /_/ / /_/ /  __/ /    
/_/    \__, /     /_____/_/ /_/\___/\____/\__,_/\___/_/     
      /____/                                                
    """,
        style="cyan",
    )
    # blink


def processSrc(mode: str, title: str):
    try:
        mode_path = "./res/" + mode
        if not os.path.exists(mode_path):
            os.makedirs(mode_path)

        if not os.path.exists("./src"):
            os.makedirs("./src")

        table = Table(title=title, style="magenta")
        table.add_column(
            "[cyan]Option[/cyan]", justify="center", style="cyan", no_wrap=True
        )
        table.add_column(
            "[cyan]Mode[/cyan]", justify="left", style="cyan", no_wrap=True
        )

        files = []
        i = 1

        for excel in os.listdir("./src"):
            if excel.endswith(".xlsx"):
                files.append(excel)
                table.add_row(str(i), str(excel))
                i += 1

        if len(files) == 0:
            console.print("⚠️ No data found in src directory.", style="warning")
            time.sleep(5)
            sys.exit(0)

        table.add_row("A", "Select All")
        table.add_row("G", "Go Back")
        console.print("\n", table, "\n")

        return files
    except Exception as e:
        if debugMode:
            console.print(":x: Error: " + str(e), style="danger")


def isEmpty(data: any):
    if (
        (pd.isna(data))
        or (data == "")
        or (data == " ")
        or (data == "nan")
        or (data == "N/A")
    ):
        return True
    else:
        return False


def isValid(data: any):
    if (
        (pd.isna(data))
        or (data == "")
        or (data == " ")
        or (data == "Note:")
        or (data == "nan")
        or not (has_numbers(data))
    ):
        return False
    else:
        return True


def clear():

    # for windows
    if name == "nt":
        _ = system("cls")

    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system("clear")


def promptExit():

    table = Table(style="magenta")
    table.add_column(
        "[cyan]Option[/cyan]", justify="center", style="cyan", no_wrap=True
    )
    table.add_column("[cyan]Mode[/cyan]", justify="left", style="cyan", no_wrap=True)

    table.add_row("C", "Continue")
    table.add_row("G", "Go Back")

    console.print("\n", table, "\n")

    opt = Prompt.ask(
        ":backhand_index_pointing_right:[blink yellow] Select an option[/blink yellow]"
    )

    if opt == "C" or opt == "c":
        return False

    else:
        return True
