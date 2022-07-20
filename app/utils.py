import shutil
import numpy as np
import pandas as pd
import sys
import time
import os
import re
import logging

from os import system, name
from os.path import exists
from app.definitions import *
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


from rich.console import Console
from rich.prompt import Prompt
from rich.theme import Theme
from rich.table import Table
from rich.progress import track


# from rich import console.print

custom_theme = Theme(
    {
        "primary": "bold cyan",
        "success": "bold bright_green",
        "info": "cyan",
        "warning": "yellow",
        "danger": "bright_red",
        "url": "not bold not italic underline bright_blue",
    }
)

console = Console(theme=custom_theme)
logger = logging.getLogger()


def enable_globalAIO():
    global is_AIO
    is_AIO = True


def disable_globalAIO():
    global is_AIO
    is_AIO = False


def updateFolderName(name: str):
    global folder_name
    folder_name = name


def resetCleanedList():
    global cleaned_log_list
    cleaned_log_list.clear()


def resetFolderName():
    global folder_name
    now = datetime.now()
    dt = now.strftime("%d%m%Y%H%M%S")
    folder_name = "Unkwnown (" + dt + ")"


def header():
    clear()

    console.print(
        r"""
    ____              ______                     __         
   / __ \__  __      / ____/___  _________  ____/ /__  _____
  / /_/ / / / /_____/ __/ / __ \/ ___/ __ \/ __  / _ \/ ___/
 / ____/ /_/ /_____/ /___/ / / / /__/ /_/ / /_/ /  __/ /    
/_/    \__, /     /_____/_/ /_/\___/\____/\__,_/\___/_/      
      /____/      [bold cyan]Version: 2.3[/bold cyan]
    """,
        style="cyan",
    )


def debugging():
    global debugMode
    debugMode = not debugMode
    return debugMode


def createLog(file_name: str, vessel: str, mode: str, desc: str):
    try:
        global folder_name
        global is_AIO
        if is_AIO:
            temp = mode.split("_")
            first = temp[0][0].upper()
            second = temp[1][0].upper()
            _mode = first + second

            creation_name = "/" + folder_name + " (" + _mode + " - Log)" + ".xlsx"
            creation_path = "./res/AIO/" + folder_name + "/" + mode
        else:
            creation_name = (
                "/" + str(file_name[: len(file_name) - 5]).strip() + " (Log)" + ".xlsx"
            )
            creation_path = "./res/" + vessel + "/" + mode + "/log/"

        if not os.path.exists(creation_path):
            os.makedirs(creation_path)

        if not exists(creation_path + creation_name):
            writer = pd.ExcelWriter(creation_path + creation_name, engine="xlsxwriter")
            writer.save()

        book = load_workbook(creation_path + creation_name)
        sheet = book.active

        global cleaned_log_list
        if is_AIO:
            if folder_name not in cleaned_log_list:
                sheet.delete_rows(1, sheet.max_row + 1)
                cleaned_log_list.append(folder_name)
        else:
            if file_name not in cleaned_log_list:
                sheet.delete_rows(1, sheet.max_row + 1)
                cleaned_log_list.append(file_name)

        rowData = (desc,)
        sheet.append(rowData)
        book.save(creation_path + creation_name)

        if debugMode:
            console.print("\n" + desc, style="danger", highlight=False)
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getFormattedDate(
    key: str,
    code: str,
    mode: str,
    file_name: str,
    date: str,
    datetype: str,
    vessel: str,
):
    if not isEmpty(date):
        date = date.strip()

    if date == "19-cot-2019":
        return "19-Oct-19"

    if date == "20-cot-2019":
        return "20-Oct-19"

    if date == "10-FE4B-22":
        return "10-Feb-22"

    if date == "2022-51":
        return "01-May-22"

    if date == "12/23/202":
        return "23-Dec-20"

    if date == "7103/2022":
        return "10-Jul-22"

    if "/" in date:
        if date.count("/") == 2 and len(date) >= 8 and len(date) <= 10:
            split_date = date.split("/")
            day = str(split_date[1])
            month = str(months[int(split_date[0]) - 1])
            year = str(split_date[2][2:])
            if len(year) == 2:
                return str(day + "-" + month + "-" + year)

        # ❌ Commissioning date "11/282021" is invalid (File: 20220403GLZ.xlsx, Sheet: Auto Pilot, Code: AP-015) : Fixed
        if date.count("/") == 1 and len(date) >= 8 and len(date) <= 10:
            split_date = date.split("/")
            day = str(split_date[1][:2])
            month = str(months[int(split_date[0]) - 1])
            year = str(split_date[1][2:][2:])
            if (
                len(year) == 2
                and (day >= 1 and day <= 31)
                and (month >= 1 and month <= 12)
            ):
                return str(day + "-" + month + "-" + year)

    createLog(
        file_name,
        vessel,
        mode,
        "❌ "
        + datetype
        + ' "'
        + date
        + '" is invalid '
        + "(File: "
        + file_name
        + ", Sheet: "
        + key
        + ", Code: "
        + code
        + ")",
    )

    return ""


def getVessels():
    try:
        vessels: list = []

        path = "./data/vessel_list.xlsx"
        ves_list = pd.read_excel(path)

        for i in range(len(ves_list.index)):
            vessels.append([str(ves_list.iloc[i, 0]), str(ves_list.iloc[i, 1])])

        return vessels
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getVessel(
    vessel_id: str,
    mode: str,
    file_name: str,
    vessels: list,
):
    try:
        if not isEmpty(vessel_id):
            vessel_id = vessel_id.strip()

        for vessel in vessels:
            if vessel[1] == vessel_id:
                return str(vessel[0])

        createLog(
            file_name,
            vessel_id,
            mode,
            "❌ No vessel found " + "(File: " + file_name + ")",
        )

        return ""

    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getMachineries():
    try:
        machineries: list = []

        path = "./data/name_list.xlsx"
        mach_list = pd.read_excel(path)

        for i in range(len(mach_list.index)):
            machineries.append([str(mach_list.iloc[i, 0]), str(mach_list.iloc[i, 1])])

        return machineries
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getMachinery(
    machinery_id: str,
    key: str,
    mode: str,
    file_name: str,
    machineries: list,
    vessel: str,
):
    try:
        if not isEmpty(machinery_id):
            machinery_id = machinery_id.strip()

        for machinery in machineries:
            if machinery[1] == machinery_id or machinery[1] == key:
                return str(machinery[0])

        createLog(
            file_name,
            vessel,
            mode,
            "❌ No machinery name found "
            + "(File: "
            + file_name
            + ", Sheet: "
            + key
            + ")",
        )

        return ""

    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getCodes():
    try:
        codes: list = []

        path = "./data/code_list.xlsx"
        code_list = pd.read_excel(path)

        for i in range(len(code_list.index)):
            codes.append([str(code_list.iloc[i, 0]), str(code_list.iloc[i, 1])])

        return codes
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getCode(
    machinery_name: str,
    key: str,
    mode: str,
    file_name: str,
    codes: list,
    vessel: str,
):
    try:
        if not isEmpty(machinery_name):
            machinery_name = machinery_name.strip()

        for code in codes:
            if code[1] == machinery_name or code[1] == key:
                return str(code[0])

        createLog(
            file_name,
            vessel,
            mode,
            "❌ No machinery code found "
            + "(File: "
            + file_name
            + ", Sheet: "
            + key
            + ", Machinery: "
            + machinery_name
            + ")",
        )

        return ""
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getIntervals():
    try:
        intervals: list = []

        path = "./data/interval_list.xlsx"
        interval_list = pd.read_excel(path)

        for i in range(len(interval_list.index)):
            intervals.append(
                [str(interval_list.iloc[i, 0]), str(interval_list.iloc[i, 1])]
            )

        return intervals
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getInterval(
    interval_id: str,
    key: str,
    mode: str,
    file_name: str,
    intervals: list,
    code: str,
    vessel: str,
):
    try:
        if not isEmpty(interval_id):
            interval_id = interval_id.strip()

        for interval in intervals:
            if interval[1] == interval_id or interval[1] == key:
                return str(interval[0])

        createLog(
            file_name,
            vessel,
            mode,
            '❌ No interval "'
            + interval_id
            + '" found '
            + "(File: "
            + file_name
            + ", Sheet: "
            + key
            + ", Code: "
            + code
            + ")",
        )

        return ""
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getIncharges(path: str):
    try:
        incharges: list = []

        main_menu = pd.read_excel(path, sheet_name="Main Menu")

        i = 1
        while i < len(main_menu.index):
            incharges.append(
                [
                    str(main_menu.iloc[i, 2]).strip(),
                    str(main_menu.iloc[i, 1]).strip(),
                ]
            )
            i += 1

        return incharges
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getIncharge(
    machinery: str,
    key: str,
    mode: str,
    file_name: str,
    incharges: list,
    vessel: str,
):
    try:
        if not isEmpty(key):
            key = key.strip()

        for incharge in incharges:
            if incharge[1] == machinery or incharge[1] == key:
                return str(incharge[0])

        createLog(
            file_name,
            vessel,
            mode,
            "❌ No incharge found "
            + "(File: "
            + file_name
            + ", Sheet: "
            + key
            + ", Machinery: "
            + machinery
            + ")",
        )

        return ""
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def saveExcelFile(book: Workbook, _filename: str, creation_folder: str):
    try:
        if not os.path.exists(creation_folder):
            os.makedirs(creation_folder)
        book.save(creation_folder + _filename)
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def hasNumbers(inputString: str):
    try:
        return bool(re.search(r"\d", inputString))
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def cleanResLog():
    try:
        isClean = False

        if os.path.exists("./log"):
            shutil.rmtree("./log")
            isClean = True

        if os.path.exists("./res"):
            shutil.rmtree("./res")
            isClean = True

        return isClean
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def emptySrc():
    try:
        if os.path.exists("./src"):
            for file in os.scandir("./src"):
                os.remove(file.path)
            return True
        return False
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def mainMenu():
    table = Table(title="[yellow] Main Menu[/yellow]", style="magenta")
    table.add_column(
        "[cyan]Option[/cyan]", justify="center", style="cyan", no_wrap=True
    )
    table.add_column("[cyan]Mode[/cyan]", justify="left", style="cyan", no_wrap=True)
    table.add_column(
        "[cyan]Type  [/cyan]", justify="center", style="cyan", no_wrap=True
    )

    table.add_row("R", "Running Hours", "🏃")
    table.add_row("S", "Sub Categories", "📚")
    table.add_row("U", "Update Jobs", "📝")
    table.add_row("V", "Vessel Machineries", "⚓")
    table.add_row("------", "------------------", "-------")
    table.add_row("A", "All-in-One", "💯")
    table.add_row("C", "Clean Res & Log", "🧹")
    table.add_row("E", "Empty Src Folder", "📂")
    table.add_row("------", "------------------", "-------")

    if debugMode:
        table.add_row("D", "Disable Debug Mode", "💻")
    else:
        table.add_row("D", "Enable Debug Mode", "💻")
    table.add_row("P", "Version History", "🕓")
    table.add_row("X", "Exit", "❌")

    console.print("", table, "\n")


def getSheetNames(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def processSrc(title: str, showExtraMenu: bool):
    try:
        header()

        table = Table(title=title, style="magenta")
        table.add_column(
            "[cyan]Option[/cyan]", justify="center", style="cyan", no_wrap=True
        )
        table.add_column(
            "[cyan]Mode[/cyan]", justify="left", style="cyan", no_wrap=True
        )
        table.add_column(
            "[cyan]Type[/cyan]", justify="left", style="cyan", no_wrap=True
        )

        files = []
        i = 1
        max_mode_length = 0

        console.print("\n")
        for excel in track(
            os.listdir("./src"), description="🟢 [bold green]Loading Files[/bold green]"
        ):
            if excel.endswith(".xlsx"):
                keys = getSheetNames("./src/" + excel)

                if "Hatch Cover" in keys:
                    table.add_row(str(i), excel, "Deck")
                    file_type = "deck"
                elif "Running Hours" in keys:
                    table.add_row(str(i), excel, "Engine")
                    file_type = "engine"
                else:
                    table.add_row(str(i), excel, "Other")
                    file_type = "other"

                max_mode_length = max(max_mode_length, len(excel))

                files.append({"type": file_type, "keys": keys, "excelFile": excel})
                i += 1

        if len(files) == 0:
            console.print(
                "\n\n⚠️ No data found in src directory.\n\n",
                style="warning",
                highlight=False,
            )
            time.sleep(10)
            sys.exit(0)

        if showExtraMenu:
            table.add_row("------", "-" * max_mode_length, "------")
            table.add_row("A", "Select All", "  💯")
            table.add_row("D", "Select Deck Only", "  ⚓")
            table.add_row("E", "Select Engine Only", "  🤖")
            table.add_row("G", "Go Back", "  🔙")
            table.add_row("R", "Refresh", "  🔃")

        return {"files": files, "table": table}

    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def isEmpty(data: any):
    try:
        if (
            (pd.isna(data))
            or (str(data).strip() == "")
            or (str(data).strip().lower() == "nan")
            or (str(data).strip().lower() == "n/a")
            or (str(data).strip().lower() == "nil")
        ):
            return True
        else:
            return False
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def isValid(data: any):
    try:
        if (
            (pd.isna(data))
            or (str(data).strip() == "")
            or (str(data).strip().lower() == "note:")
            or (str(data).strip().lower() == "nan")
            or (str(data).strip().lower() == "n/a")
            or (str(data).strip().lower() == "nil")
            or not (hasNumbers(str(data)))
        ):
            return False
        else:
            return True
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def clear():
    # for windows
    if name == "nt":
        _ = system("cls")
    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system("clear")


def showExitCredits():
    header()

    console.print(
        "\n\n💻 Source: " + "[url]https://github.com/vn-aj-vngrd/py-encoder[/url]"
    )
    console.print("💛 Created by: " + "[warning]Van AJ B. Vanguardia[/warning]\n\n")

    for _ in track(range(100), description="[bright_red]🔴 Exiting[/bright_red]\n\n"):
        time.sleep(0.01)

    sys.exit(0)


def promptExitorContinue():

    table = Table(style="magenta")
    table.add_column(
        "[cyan]Option[/cyan]", justify="center", style="cyan", no_wrap=True
    )
    table.add_column("[cyan]Mode[/cyan]", justify="left", style="cyan", no_wrap=True)

    table.add_row("C", "Continue")
    table.add_row("G", "Go Back")

    console.print("\n", table, "\n")

    opt = Prompt.ask(
        ":backhand_index_pointing_right:[blink yellow] Select an option[/blink yellow]"
    )

    if opt == "C" or opt == "c":
        return False

    else:
        return True


def promptExit():
    table = Table(style="magenta")
    table.add_column(
        "[cyan]Option[/cyan]", justify="center", style="cyan", no_wrap=True
    )
    table.add_column("[cyan]Mode[/cyan]", justify="left", style="cyan", no_wrap=True)

    table.add_row("G", "Go Back")
    console.print("\n", table, "\n")

    _ = Prompt.ask(
        ":backhand_index_pointing_right:[blink yellow] Select an option[/blink yellow]"
    )


def isFloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def displayVersionHistory():
    header()

    console.print("\n🕓 Version History\n\n", style="warning")

    for ver in version_history:
        console.print(ver)

    promptExit()


def splitAIO(_dir: str, file_name: str, mode: str, chunksize: int):
    try:
        df = pd.read_excel(_dir + file_name)
        global folder_name

        i = 0
        for chunk in track(
            np.array_split(df, len(df) // chunksize),
            description="🟢 [bold green]Splitting  [/bold green]",
        ):
            chunk.to_excel(
                # _dir + "/AIO_{:02d}".format(i) + " (" + mode + ").xlsx",
                _dir + "/" + folder_name + "_{:02d}".format(i) + " (" + mode + ").xlsx",
                index=False,
                header=True,
            )
            i += 1

        return i
    except Exception as e:
        if debugMode:
            logger.exception(e, stack_info=True)


def getMinVal(row_count: str):
    spilt_n = 2
    global base

    if row_count >= base:
        i = spilt_n
        while True:
            if (base * i) >= row_count:
                return row_count / i
            i += 1

    return row_count
