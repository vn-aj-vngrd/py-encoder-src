from os.path import exists
from app.definitions import *
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import os
import re


def getMachinery(machineryCode, key, mode, file_name):
    try:
        path = "./app/data/gen_mach_list.xlsx"
        mach_list = pd.read_excel(path)

        i = 0
        while (not pd.isna(mach_list.iloc[i, 1])) and (
            mach_list.iloc[i, 1] != machineryCode
        ):
            i += 1
            if mach_list.iloc[i, 1] == "ECT":
                break

        if not pd.isna(mach_list.iloc[i, 1]) and (
            mach_list.iloc[i, 1] == machineryCode
        ):
            return mach_list.iloc[i, 0]
        else:
            creation_name = "/" + file_name
            creation_path = "./bin/" + mode

            if not os.path.exists(creation_path):
                os.makedirs(creation_path)

            if not exists(creation_path + creation_name):
                writer = pd.ExcelWriter(
                    creation_path + creation_name, engine="xlsxwriter"
                )
                writer.save()
                book = load_workbook(creation_path + creation_name)
                sheet = book.active
                sheet.append(bin_header)
                book.save(creation_path + creation_name)

            book = load_workbook(creation_path + creation_name)
            sheet = book.active

            rowData = (key, machineryCode)
            sheet.append(rowData)
            book.save(creation_path + creation_name)

            print(
                "\nError: No such machinery code found for "
                + key
                + ": "
                + machineryCode
                + "\n"
            )
            return "N/A"

    except Exception as e:
        print("Error: " + str(e) + " (" + key + ": " + machineryCode + ")" + "\n")


def main_function():
    if not os.path.exists("./res/main"):
        os.makedirs("./res/main")

    while True:
        if not os.path.exists("./src"):
            os.makedirs("./src")

        files = []
        i = 0
        for excel in os.listdir("./src"):
            if excel.endswith(".xlsx"):
                files.append(excel)
                print(i, "-", excel)
                i += 1

        if len(files) == 0:
            print("No such data found in src directory.")
            exit()

        # Get the location of the data
        try:
            file_key = input("\nInput file number: ")
            file_name = files[int(file_key)]
            path = "src/" + file_name
            print("Excel File: " + file_name)
        except Exception as e:
            print("Error: ", str(e))

        # Read the data
        try:
            data = pd.read_excel(path, sheet_name=None, index_col=None, header=None)

            # Get the keys
            xl = pd.ExcelFile(path)
            keys = xl.sheet_names

            # Iterate through the sheets
            for key in keys:
                if key not in notIncluded:
                    print(key)

                    # Vessel Name
                    vessel = data[key].iloc[0, 2]

                    # Default Machinery Name: machinery = data[key].iloc[2, 2]
                    # Machinery Name using the machinery code
                    machinery = getMachinery(
                        str(data[key].iloc[2, 5]), key, "main", file_name
                    )

                    # Start traversing the data on row 7
                    row = 7
                    isValid = True

                    # Prepare the sheets
                    book = Workbook()
                    sheet = book.active

                    sheet.append(main_header)

                    while isValid:

                        rowData = (
                            vessel,
                            machinery,
                        )

                        for col in range(7):
                            d = data[key].iloc[row, col]

                            if (pd.isna(d)) and (col == 0):
                                isValid = False
                                break

                            if pd.isna(d):
                                d = " "

                            if (col == 3) and not (re.search("[a-zA-Z]", str(d))):
                                d = str(d) + " Hours"

                            if ((col == 4) or (col == 5)) and isinstance(d, datetime):
                                d = d.strftime("%d-%b-%y")
                            else:
                                d = re.sub("\\s+", " ", str(d))

                            tempTuple = (d,)
                            rowData += tempTuple

                        if isValid:
                            sheet.append(rowData)
                            row += 1

                    create_name = file_name[: len(file_name) - 4]
                    creation_folder = "./res/main/" + create_name
                    if not os.path.exists(creation_folder):
                        os.makedirs(creation_folder)
                    book.save(creation_folder + "/" + key + ".xlsx")

            print("Done...")
        except Exception as e:
            print("Error: " + str(e))

        isContinue = input("Input 1 to continue: ")
        if isContinue != "1":
            break


def sub_function():
    if not os.path.exists("./res/sub"):
        os.makedirs("./res/sub")

    while True:
        if not os.path.exists("./src"):
            os.makedirs("./src")

        files = []
        i = 0
        for excel in os.listdir("./src"):
            if excel.endswith(".xlsx"):
                files.append(excel)
                print(i, "-", excel)
                i += 1

        if len(files) == 0:
            print("No such data found in src directory.")
            exit()

        # Get the location of the data
        try:
            file_key = input("\nInput file number: ")
            file_name = files[int(file_key)]
            path = "src/" + file_name
            print("Excel File: " + file_name)
        except Exception as e:
            print("Error: ", str(e))

        # Read the data
        try:
            data = pd.read_excel(path, sheet_name=None, index_col=None, header=None)

            # Get the keys
            xl = pd.ExcelFile(path)
            keys = xl.sheet_names

            # Get updated_at
            updating_date = data["Running Hours"].iloc[2, 3].strftime("%d-%b-%y")
            # print(updating_date)

            # Prepare the sheets
            book = Workbook()
            sheet = book.active

            # Append the dates
            sheet.append(sub_header)

            # Iterate through the sheets
            for key in keys:
                if key not in notIncluded:
                    print(key)

                    # Vessel Name
                    vessel = data[key].iloc[0, 2]

                    # Default Machinery Name: machinery = data[key].iloc[2, 2]
                    # Machinery Name
                    machinery = getMachinery(
                        str(data[key].iloc[2, 5]), key, "sub", file_name
                    )

                    # Running Hours
                    running_hours = data[key].iloc[3, 5]

                    rowData = (vessel, machinery, running_hours, updating_date)
                    sheet.append(rowData)

            create_name = file_name[: len(file_name) - 4]
            creation_folder = "./res/sub/" + create_name
            if not os.path.exists(creation_folder):
                os.makedirs(creation_folder)
            book.save(creation_folder + "/" + file_name)

            print("Done...")
        except Exception as e:
            print("Error: " + str(e))

        isContinue = input("Input 1 to continue: ")
        if isContinue != "1":
            break
